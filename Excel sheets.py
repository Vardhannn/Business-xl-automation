import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def work_book(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    Values = Reference(sheet, min_row=2, max_row=4, min_col=3, max_col=3)

    chart = BarChart()
    chart.add_data(Values)
    sheet.add_chart(chart, "a7")

    wb.save(filename)
